# 셀 병합

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2")            # B2 부터 D2 까지 셀 병합
ws["B2"].value = "Merged Cell"

wb.save("sample_merge.xlsx")